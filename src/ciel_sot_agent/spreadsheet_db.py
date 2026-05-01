"""CIEL Spreadsheet DB — arkusze kalkulacyjne jako baza danych kart obiektów.

Każda kategoria kart to osobny arkusz w jednym pliku XLSX:
  integration/db/ciel_cards.xlsx

Arkusze:
  entity_cards   — karty bytów (id, noun, coupling, phase, horizon_class, adjectives, note)
  htri_metrics   — historia metryk HTRI (timestamp, coherence, n_threads, kappa)
  pipeline_metrics — historia pipeline (timestamp, cycle, ethical, soul, mood, closure, emotion)
  cqcl_log       — log CQCL per wywołanie (timestamp, input_hash, coherence, emotion, intensity)
  nonlocal_cards — karty nielokalności (id, type, description, active)

Zasada: APPEND ONLY na arkuszach *_metrics i *_log.
Karty obiektów (entity_cards, nonlocal_cards) są upsertowane po id.
"""
from __future__ import annotations

import fcntl
import hashlib
import time
from contextlib import contextmanager
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

_DB_PATH = Path(__file__).parents[2] / "integration" / "db" / "ciel_cards.xlsx"
_LOCK_PATH = _DB_PATH.with_suffix(".lock")


@contextmanager
def _xlsx_lock():
    _LOCK_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(_LOCK_PATH, "w") as lf:
        fcntl.flock(lf, fcntl.LOCK_EX)
        try:
            yield
        finally:
            fcntl.flock(lf, fcntl.LOCK_UN)

# ── Schematy nagłówków ────────────────────────────────────────────────────────

_SCHEMAS: dict[str, list[str]] = {
    "entity_cards": [
        "id", "noun", "coupling_ciel", "phase", "horizon_class",
        "adjectives", "note", "last_updated",
    ],
    "htri_metrics": [
        "timestamp", "datetime", "coherence", "n_threads_optimal",
        "kappa", "n_oscillators",
    ],
    "pipeline_metrics": [
        "timestamp", "datetime", "cycle", "ethical_score", "soul_invariant",
        "mood", "closure_penalty", "dominant_emotion", "system_health",
        "htri_coherence",
    ],
    "cqcl_log": [
        "timestamp", "datetime", "input_hash", "input_preview",
        "quantum_coherence", "dominant_emotion", "emotional_intensity",
        "htri_r", "bridge_active",
    ],
    "nonlocal_cards": [
        "id", "type", "description", "active", "coupling", "last_updated",
    ],
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _now_ts() -> tuple[float, str]:
    now = datetime.now(timezone.utc)
    return now.timestamp(), now.strftime("%Y-%m-%dT%H:%M:%SZ")


def _load_or_create() -> Workbook:
    _DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    if _DB_PATH.exists():
        return load_workbook(str(_DB_PATH))
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, headers in _SCHEMAS.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        # Pogrubienie nagłówka
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
    wb.save(str(_DB_PATH))
    return wb


def _ensure_sheet(wb: Workbook, name: str) -> Worksheet:
    if name not in wb.sheetnames:
        ws = wb.create_sheet(name)
        ws.append(_SCHEMAS.get(name, ["value"]))
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
    return wb[name]


def _row_by_id(ws: Worksheet, id_val: str, id_col: int = 1) -> int | None:
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[id_col - 1].value == id_val:
            return row[id_col - 1].row
    return None


# ── Public API ────────────────────────────────────────────────────────────────

def upsert_entity_card(entity: dict[str, Any]) -> None:
    """Upsert karty bytu po id. Tworzy lub aktualizuje wiersz."""
    with _xlsx_lock():
        wb = _load_or_create()
        ws = _ensure_sheet(wb, "entity_cards")
        _, dt = _now_ts()
        adj = "|".join(entity.get("adjectives", []))
        row_data = [
            entity.get("id", ""),
            entity.get("noun", ""),
            entity.get("coupling_ciel", 0.0),
            entity.get("phase", 0.0),
            entity.get("horizon_class", ""),
            adj,
            entity.get("note", ""),
            dt,
        ]
        existing = _row_by_id(ws, entity.get("id", ""))
        if existing:
            for col, val in enumerate(row_data, 1):
                ws.cell(row=existing, column=col, value=val)
        else:
            ws.append(row_data)
        wb.save(str(_DB_PATH))


def append_htri_metrics(state: dict[str, Any]) -> None:
    """Dopisz pomiar HTRI do arkusza htri_metrics."""
    with _xlsx_lock():
        wb = _load_or_create()
        ws = _ensure_sheet(wb, "htri_metrics")
        ts, dt = _now_ts()
        ws.append([
            ts, dt,
            state.get("coherence", 0.0),
            state.get("n_threads_optimal", 0),
            state.get("kappa", 0.0),
            state.get("n_oscillators", 12),
        ])
        wb.save(str(_DB_PATH))


def append_pipeline_metrics(report: dict[str, Any]) -> None:
    """Dopisz wynik pipeline do arkusza pipeline_metrics."""
    with _xlsx_lock():
        wb = _load_or_create()
        ws = _ensure_sheet(wb, "pipeline_metrics")
        ts, dt = _now_ts()
        ws.append([
            ts, dt,
            report.get("cycle", 0),
            report.get("ethical_score", 0.0),
            report.get("soul_invariant", 0.0),
            report.get("mood", 0.0),
            report.get("closure_penalty", 0.0),
            report.get("dominant_emotion", ""),
            report.get("system_health", 0.0),
            report.get("htri_coherence", 0.0),
        ])
        wb.save(str(_DB_PATH))


def append_cqcl_log(
    cqcl_input: str,
    metrics: dict[str, Any],
    htri_r: float = 0.85,
    bridge_active: bool = False,
) -> None:
    """Dopisz log wywołania CQCL."""
    with _xlsx_lock():
        wb = _load_or_create()
        ws = _ensure_sheet(wb, "cqcl_log")
        ts, dt = _now_ts()
        h = hashlib.md5(cqcl_input.encode()).hexdigest()[:8]
        ws.append([
            ts, dt, h,
            cqcl_input[:60],
            metrics.get("quantum_coherence", 0.0),
            metrics.get("dominant_emotion", ""),
            metrics.get("emotional_intensity", 0.0),
            htri_r,
            bridge_active,
        ])
        wb.save(str(_DB_PATH))


def upsert_nonlocal_card(card: dict[str, Any]) -> None:
    """Upsert karty nielokalności po id."""
    with _xlsx_lock():
        wb = _load_or_create()
        ws = _ensure_sheet(wb, "nonlocal_cards")
        _, dt = _now_ts()
        row_data = [
            card.get("id", ""),
            card.get("type", ""),
            card.get("description", ""),
            card.get("active", True),
            card.get("coupling", 0.0),
            dt,
        ]
        existing = _row_by_id(ws, card.get("id", ""))
        if existing:
            for col, val in enumerate(row_data, 1):
                ws.cell(row=existing, column=col, value=val)
        else:
            ws.append(row_data)
        wb.save(str(_DB_PATH))


def load_entity_cards() -> list[dict[str, Any]]:
    """Wczytaj wszystkie karty bytów z arkusza."""
    if not _DB_PATH.exists():
        return []
    wb = load_workbook(str(_DB_PATH), read_only=True)
    if "entity_cards" not in wb.sheetnames:
        return []
    ws = wb["entity_cards"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    cards = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        d = dict(zip(headers, row))
        if d.get("adjectives"):
            d["adjectives"] = str(d["adjectives"]).split("|")
        cards.append(d)
    return cards


def db_path() -> Path:
    return _DB_PATH
